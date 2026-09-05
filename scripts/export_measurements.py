"""Measure every labelled specimen in a dataset and write the Excel workbook.

A thin wrapper over ``fish_morpho.pipeline`` that knows the repository's layout,
so producing the spreadsheet is one command with a dataset name rather than three
paths that have to agree with each other.

    python scripts/export_measurements.py --dataset alewife

Writes ``results/<dataset>/measurements.xlsx`` with six sheets:

  About         what the file is, the commit it came from, the millimetre /
                pixel split, and the check counts — so the workbook can be read
                by someone who did not run it.
  Measurements  one row per specimen, one column per trait, plus a ``units``
                column — a series photographed without a usable scale reference
                measures in PIXELS, and mixing those with millimetres under one
                header is a mistake nothing downstream can catch.
  Ratios        every length divided by standard length and every area by SL
                squared, so the numbers are dimensionless and comparable between
                fish of different sizes. This is the sheet a between-population
                comparison wants, and the only meaningful one when there is no
                scale.
  Shape         Mosimann log-shape variables: each length over the geometric
                mean of all lengths, logged. The defensible size correction when
                the groups being compared may differ in body size.
  QC            calibration method and confidence per view, which landmarks were
                missing, and any recorded data compromise.
  Validation    the automated checks, most severe first. Read before analysing.

Specimens that cannot be processed are named and skipped rather than aborting the
batch.
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT / "src"))

from fish_morpho.landmark_config import traits_requiring  # noqa: E402
from fish_morpho.pipeline import run  # noqa: E402


def dropped_traits(dataset_dir: Path) -> tuple[str, ...]:
    """Traits the dataset's schema profile puts out of scope.

    A landmark the study never collects would otherwise leave a column of blanks,
    which reads as "measured and missing" rather than "never in scope".
    """
    f = dataset_dir / "schema.json"
    if not f.is_file():
        return ()
    try:
        prof = json.loads(f.read_text())
    except Exception:
        return ()
    return tuple(traits_requiring(prof.get("exclude_keypoints") or [],
                                  prof.get("exclude_polygons") or []))

log = logging.getLogger("export_measurements")


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(prog="export_measurements")
    ap.add_argument("--dataset", help="Folder name under --data-root, e.g. alewife.")
    ap.add_argument("--data-root", type=Path, default=_ROOT / "data")
    ap.add_argument("--images", type=Path, default=None,
                    help="Override the image directory.")
    ap.add_argument("--labels", type=Path, default=None,
                    help="Override the sidecar directory.")
    ap.add_argument("--out", type=Path, default=None)
    ap.add_argument("--log-level", default="WARNING",
                    choices=("DEBUG", "INFO", "WARNING", "ERROR"))
    args = ap.parse_args(argv)

    logging.basicConfig(level=getattr(logging, args.log_level),
                        format="%(levelname)s %(message)s")

    if args.images and args.labels:
        images, labels = args.images, args.labels
        name = args.dataset or images.parent.name
    else:
        if not args.dataset:
            avail = sorted(d.name for d in args.data_root.iterdir()
                           if d.is_dir() and (d / "lateral").is_dir()) \
                    if args.data_root.is_dir() else []
            ap.error("--dataset is required (or pass --images and --labels). "
                     f"Available: {', '.join(avail) or 'none found'}")
        base = args.data_root / args.dataset
        images, labels, name = base / "lateral", base / "sidecars", args.dataset
        if not images.is_dir():
            ap.error(f"{images} does not exist")

    out = args.out or (_ROOT / "results" / name / "measurements.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)

    try:
        written = run(images_dir=images, labels_dir=labels, output_path=out,
                      mode="manual", model_config=None,
                      drop_traits=dropped_traits(images.parent))
    except Exception as exc:
        log.error("%s", exc)
        return 1

    from openpyxl import load_workbook
    wb = load_workbook(written)
    ws = wb["Measurements"]
    hdr = [c.value for c in ws[1]]
    u = hdr.index("units")
    units = [r[u] for r in ws.iter_rows(min_row=2, values_only=True)]
    dropped = dropped_traits(images.parent)
    print(f"wrote {written}")
    if dropped:
        print(f"  omitted {len(dropped)} trait column(s) out of scope for this "
              f"study: {', '.join(sorted(dropped))}")
    print(f"  {ws.max_row - 1} specimens x {len(hdr)} columns, "
          f"sheets: {', '.join(wb.sheetnames)}")
    if units:
        mm, px = units.count("mm"), units.count("px")
        print(f"  units: {mm} in mm, {px} in pixels"
              + ("   <- MIXED; use the Ratios sheet to compare across rows"
                 if mm and px else ""))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
