"""Excel export for a batch of measurement sets.

Six sheets, because a bare measurement table answers fewer questions than it
appears to: the numbers alone cannot say what units a row is in, how a group
comparison should be size-corrected, or whether anything about the batch looked
wrong on the way out. About, Ratios, Shape, QC and Validation each carry one of
those, so the workbook explains itself to someone who did not run it.

We use openpyxl directly (no pandas) to keep dependencies light.
"""

from __future__ import annotations

import math
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .landmark_config import Unit
from .measurement_engine import (
    MeasurementSet,
    measurement_column_order,
    measurement_labels,
)
from .ruler_calibration import CalibrationResult


DEFAULT_METADATA_COLUMNS: tuple[str, ...] = (
    "fish_id",
    "locality",
    "collection_date",
    "image_filename",
)


@dataclass
class ExportRecord:
    """One specimen's data packaged for export."""

    measurements: MeasurementSet
    # Calibrations per view — rendered on the QC sheet for provenance.
    calibrations: dict[str, CalibrationResult]
    image_filename: str = ""
    # Raw annotation and frame size, carried so the validation pass can check
    # things the measurements alone cannot reveal: a mirrored specimen, or a
    # sidecar written against a different photograph.
    keypoints: dict[str, tuple[float, float]] = field(default_factory=dict)
    polygons: dict[str, list] = field(default_factory=dict)
    image_size: tuple[int, int] | None = None


def export_to_xlsx(
    records: Sequence[ExportRecord],
    output_path: str | Path,
    metadata_columns: Iterable[str] = DEFAULT_METADATA_COLUMNS,
    drop_traits: Iterable[str] = (),
    issues: Sequence = None,
    provenance: dict | None = None,
) -> Path:
    """Write ``records`` to an xlsx workbook at ``output_path``.

    The workbook has six sheets:

    * ``About`` — what the file is, when and from which commit it came, the
      unit split, the check counts, and any traits held out of scope.
    * ``Measurements`` — metadata columns + one column per measurement,
      with numeric values in mm / mm^2.
    * ``Ratios`` — every length over standard length and every area over SL
      squared, so the values are dimensionless and comparable between fish of
      different sizes. The only sheet where a scale-free specimen and a
      calibrated one can honestly share a column.
    * ``Shape`` — Mosimann log-shape variables, each length divided by the
      geometric mean of all lengths and logged. The size correction to use when
      comparing groups that may differ in size; see ``_write_shape_sheet``.
    * ``QC`` — calibration method / confidence / notes per view, plus a
      ``missing_landmarks`` column summarizing any gaps.
    * ``Validation`` — the checks from :mod:`fish_morpho.validation`, most
      severe first. Written only when ``issues`` is passed.

    Returns the resolved path.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    meas_sheet = wb.active
    assert meas_sheet is not None
    meas_sheet.title = "Measurements"
    drop = set(drop_traits)
    _write_measurements_sheet(meas_sheet, records, list(metadata_columns), drop)

    ratio_sheet = wb.create_sheet("Ratios")
    _write_ratios_sheet(ratio_sheet, records, list(metadata_columns), drop)

    shape_sheet = wb.create_sheet("Shape")
    _write_shape_sheet(shape_sheet, records, list(metadata_columns), drop)

    qc_sheet = wb.create_sheet("QC")
    _write_qc_sheet(qc_sheet, records)

    if issues is not None:
        _write_validation_sheet(wb.create_sheet("Validation"), issues)

    # Last so it lands rightmost, first so it is what opens: moved to index 0.
    about = wb.create_sheet("About")
    _write_about_sheet(about, records, drop, issues, provenance or {})
    wb.move_sheet("About", offset=-(len(wb.sheetnames) - 1))
    wb.active = 0

    wb.save(output_path)
    return output_path


def _write_measurements_sheet(
    sheet: Worksheet,
    records: Sequence[ExportRecord],
    metadata_columns: list[str],
    drop_traits: set[str] = frozenset(),
) -> None:
    measurement_keys = [k for k in measurement_column_order() if k not in drop_traits]
    labels = measurement_labels()

    # A workbook can hold both scaled and scale-free specimens -- a series shot
    # without a usable ruler measures in PIXELS, and putting those under an "(mm)"
    # header beside real millimetres is the kind of mistake nothing downstream can
    # catch. The units of each row travel with the row.
    header = [*metadata_columns, "units", *(labels[k] for k in measurement_keys)]
    sheet.append(header)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E6E6E6")
    for col_idx in range(1, len(header) + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for rec in records:
        row: list[float | str] = []
        for col in metadata_columns:
            if col == "image_filename":
                row.append(rec.image_filename)
            elif col == "fish_id":
                row.append(rec.measurements.fish_id)
            else:
                row.append(rec.measurements.metadata.get(col, ""))
        # Three states, not two. A specimen with no lateral calibration at all is
        # not "in pixels" -- it has no lateral measurements to have units for, and
        # saying px would invent a claim about empty cells.
        lat = rec.calibrations.get("lateral")
        row.append("" if lat is None else ("px" if lat.method == "none" else "mm"))
        for key in measurement_keys:
            v = rec.measurements.values.get(key)
            if v is None or math.isnan(v.value):
                row.append("")
            else:
                row.append(round(v.value, 3))
        sheet.append(row)

    # Reasonable column widths.
    for col_idx in range(1, len(header) + 1):
        letter = sheet.cell(row=1, column=col_idx).column_letter
        sheet.column_dimensions[letter].width = max(
            14, min(40, len(str(header[col_idx - 1])) + 2)
        )


def _write_ratios_sheet(
    sheet: Worksheet,
    records: Sequence[ExportRecord],
    metadata_columns: list[str],
    drop_traits: set[str] = frozenset(),
) -> None:
    """Size-corrected traits: lengths / SL, areas / SL^2, angles unchanged.

    Comparing raw lengths between populations mostly compares how big the fish
    happened to be, so a shape comparison wants ratios; forming them here means
    everyone forms them the same way.

    Ratios also need no calibration. Every landmark on a planar specimen shares
    one magnification, so trait/SL is exact whatever that magnification is --
    which makes this the only sheet where a scale-free specimen and a calibrated
    one can honestly sit in the same column.
    """
    keys = [k for k in measurement_column_order() if k not in drop_traits]
    labels = measurement_labels()

    header = [*metadata_columns, *(f"{labels[k]} /SL" for k in keys)]
    sheet.append(header)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E6E6E6")
    for col_idx in range(1, len(header) + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for rec in records:
        sl_val = rec.measurements.values.get("SL")
        sl = sl_val.value if sl_val is not None else float("nan")
        row: list[float | str] = []
        for col in metadata_columns:
            if col == "image_filename":
                row.append(rec.image_filename)
            elif col == "fish_id":
                row.append(rec.measurements.fish_id)
            else:
                row.append(rec.measurements.metadata.get(col, ""))
        for key in keys:
            v = rec.measurements.values.get(key)
            if v is None or math.isnan(v.value):
                row.append("")
            elif v.unit == Unit.DEG:            # an angle is already scale-free
                row.append(round(v.value, 3))
            elif math.isnan(sl) or sl <= 0:
                row.append("")                  # no SL, no ratio
            elif v.unit == Unit.MM2:
                row.append(round(v.value / (sl ** 2), 6))
            else:
                row.append(round(v.value / sl, 6))
        sheet.append(row)

    for col_idx in range(1, len(header) + 1):
        letter = sheet.cell(row=1, column=col_idx).column_letter
        sheet.column_dimensions[letter].width = max(
            14, min(40, len(str(header[col_idx - 1])) + 2))


def _write_validation_sheet(sheet: Worksheet, issues: Sequence) -> None:
    """Every check that fired, most severe first.

    In the workbook rather than only in a terminal, because the workbook is what
    gets emailed, opened months later, and handed to someone who never ran the
    export. A caveat that lives in a console scrollback is a caveat nobody has.
    """
    sheet.append(["level", "check", "fish_id", "detail"])
    for i, c in enumerate(sheet[1], start=1):
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="E6E6E6")
    colour = {"error": "FFC7CE", "warning": "FFEB9C", "note": "EDEDED"}
    for iss in issues:
        sheet.append([iss.level, iss.check, iss.fish_id, iss.message])
        fill = PatternFill("solid", fgColor=colour.get(iss.level, "FFFFFF"))
        sheet.cell(row=sheet.max_row, column=1).fill = fill
    if not issues:
        sheet.append(["", "", "", "No checks fired."])
    for col, w in zip("ABCD", (10, 22, 34, 110)):
        sheet.column_dimensions[col].width = w
    sheet.freeze_panes = "A2"


def _write_about_sheet(sheet: Worksheet, records, drop_traits, issues,
                       provenance: dict) -> None:
    """What this file is, how it was made, and what not to do with it."""
    def unit_of(rec):
        lat = rec.calibrations.get("lateral")
        return None if lat is None else ("px" if lat.method == "none" else "mm")

    mm = sum(1 for r in records if unit_of(r) == "mm")
    px = sum(1 for r in records if unit_of(r) == "px")
    counts = {"error": 0, "warning": 0, "note": 0}
    for i in (issues or []):
        counts[i.level] = counts.get(i.level, 0) + 1

    rows: list[tuple[str, str]] = [
        ("caliPr measurement export", ""),
        ("", ""),
        ("dataset", str(provenance.get("dataset", ""))),
        ("generated", provenance.get("generated", "")),
        ("caliPr commit", provenance.get("commit", "unknown")),
        ("specimens", str(len(records))),
        ("", ""),
        ("SHEETS", ""),
        ("Measurements", "One row per specimen, one column per trait. The 'units' "
                         "column says whether THAT ROW is millimetres or pixels."),
        ("Ratios", "Each length over standard length, each area over SL squared. "
                   "Dimensionless. Assumes shape does not change with size."),
        ("Shape", "Mosimann log-shape variables: each length over the geometric "
                  "mean of all lengths, logged. Use this to compare groups that "
                  "may differ in body size — it is the defensible size correction."),
        ("QC", "Calibration method and confidence per view, missing landmarks, "
               "and any recorded data compromise, per specimen."),
        ("Validation", "Automated checks. Read this before analysing."),
        ("", ""),
        ("UNITS", ""),
        ("millimetres", f"{mm} specimen(s)"),
        ("pixels (no scale reference)", f"{px} specimen(s)"),
        ("", "A specimen photographed without a usable scale measures in PIXELS. "
             "Never compare a raw length across rows of different units — use "
             "Shape or Ratios, which are unitless."
         if (mm and px) else ""),
        ("", ""),
        ("CHECKS", f"{counts['error']} error(s), {counts['warning']} warning(s), "
                   f"{counts['note']} note(s) — see the Validation sheet"),
        ("", ""),
        ("OUT OF SCOPE", ", ".join(sorted(drop_traits)) if drop_traits
                         else "no traits excluded"),
        ("", "Traits the study does not collect have no column at all, rather "
             "than a column of blanks."
         if drop_traits else ""),
    ]
    for k, v in rows:
        sheet.append([k, v])
    sheet["A1"].font = Font(bold=True, size=14)
    for r in range(1, sheet.max_row + 1):
        a = sheet.cell(row=r, column=1)
        if a.value in ("SHEETS", "UNITS", "CHECKS", "OUT OF SCOPE"):
            a.font = Font(bold=True)
        sheet.cell(row=r, column=2).alignment = Alignment(wrap_text=True,
                                                          vertical="top")
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 100


def _write_shape_sheet(
    sheet: Worksheet,
    records: Sequence[ExportRecord],
    metadata_columns: list[str],
    drop_traits: set[str] = frozenset(),
) -> None:
    """Mosimann log-shape variables: log(trait) - log(geometric mean of traits).

    The Ratios sheet divides by standard length, which is only a fair size
    correction if shape does not change with size. It usually does, and if one
    group is systematically smaller than the other -- landlocked forms often are --
    dividing by SL leaves size sitting inside the "shape" numbers and a population
    difference can be read where only a size difference exists.

    Dividing instead by the GEOMETRIC MEAN of all the length traits gives the
    standard isometric size correction. The variables are dimensionless, so a
    scale-free specimen and a calibrated one are directly comparable, and their
    log scale is what the usual multivariate tools (PCA, MANOVA) assume.

    Lengths only. An area does not share units with a length, and an angle is
    already scale-free, so folding either into the geometric mean would be
    meaningless.
    """
    keys = [k for k in measurement_column_order()
            if k not in drop_traits and _trait_unit(k) == Unit.MM]
    labels = measurement_labels()

    header = [*metadata_columns, "size (geom. mean)",
              *(f"log {labels[k].split(' — ')[0]}" for k in keys)]
    sheet.append(header)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E6E6E6")
    for col_idx in range(1, len(header) + 1):
        c = sheet.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    for rec in records:
        vals = {}
        for k in keys:
            mv = rec.measurements.values.get(k)
            if mv is not None and not math.isnan(mv.value) and mv.value > 0:
                vals[k] = mv.value
        row: list[float | str] = []
        for col in metadata_columns:
            if col == "image_filename":
                row.append(rec.image_filename)
            elif col == "fish_id":
                row.append(rec.measurements.fish_id)
            else:
                row.append(rec.measurements.metadata.get(col, ""))
        if len(vals) < 3:
            # too few traits to define a size; a two-trait geometric mean is just
            # the pair, and the "shape" would be an artefact of which two survived
            sheet.append([*row, "", *([""] * len(keys))])
            continue
        gm = math.exp(sum(math.log(v) for v in vals.values()) / len(vals))
        row.append(round(gm, 4))
        for k in keys:
            row.append(round(math.log(vals[k] / gm), 6) if k in vals else "")
        sheet.append(row)

    for col_idx in range(1, len(header) + 1):
        letter = sheet.cell(row=1, column=col_idx).column_letter
        sheet.column_dimensions[letter].width = max(
            14, min(40, len(str(header[col_idx - 1])) + 2))


def _trait_unit(code: str):
    from .landmark_config import TRAITS
    for t in TRAITS:
        if t.code == code:
            return t.unit
    return None


def _write_qc_sheet(sheet: Worksheet, records: Sequence[ExportRecord]) -> None:
    header = [
        "fish_id",
        "image_filename",
        "view",
        "calibration_method",
        "px_per_mm",
        "confidence",
        "calibration_notes",
        "missing_landmarks",
        "data_note",
    ]
    sheet.append(header)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E6E6E6")
    for col_idx in range(1, len(header) + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    for rec in records:
        missing = sorted(
            {
                lm
                for mv in rec.measurements.values.values()
                for lm in mv.missing_landmarks
            }
        )
        missing_str = ", ".join(missing) if missing else ""
        data_note = str(rec.measurements.metadata.get("data_note", "") or "")
        for view_name, calib in rec.calibrations.items():
            sheet.append(
                [
                    rec.measurements.fish_id,
                    rec.image_filename,
                    view_name,
                    calib.method,
                    round(calib.px_per_mm, 4),
                    round(calib.confidence, 3),
                    calib.notes,
                    missing_str,
                    data_note,
                ]
            )

    for col_idx in range(1, len(header) + 1):
        letter = sheet.cell(row=1, column=col_idx).column_letter
        sheet.column_dimensions[letter].width = 18
